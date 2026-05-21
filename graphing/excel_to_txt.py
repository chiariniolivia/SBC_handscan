import sys, os, glob, csv
from openpyxl import load_workbook


"""
take in excel file as arugment
output to txt file named "whateverwahtever.txt"
put header info into file
for row in excelsheet:
    if run != null and ev != null:
        add event to list of events to add




"""
scanner = "ochiarin"

header = f"Output from ped hand scanning.\nrun  ev  scanner  scan_source  scan_nbub  scan_trigger  scan_crosshairsgood  scan_comment\n%s  %d  %s  %d  %d  %d  %d  %s\n\n"




## format (run, event, mult)





path = "/home/neutron/coop/sheets/SBC_HandscanComparision.xlsx"
wb = load_workbook(filename=path, read_only=True, data_only=True)    
ws = wb.active
eventsToAdd = []
for row in ws.iter_rows(min_row=3):
    col1, col2 = row[0].value, row[1].value
    if col1 not in (None, "") and col2 not in (None, ""):
        eventsToAdd.append((col1,col2, row[2].value))



with open("scan_aaa_ochiarin_Thu_May_21_15_09_0_2026.txt",'w') as f:
    f.write(header)
    for event in eventsToAdd:
        # just put 0 for event location for now
        stringOut = f"{event[0]}  {event[1]}  {scanner}  0 {event[2]} 1 1\n"
        f.write(stringOut)


