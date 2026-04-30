import matplotlib.pyplot as plt
import numpy as np
import sys
from openpyxl import load_workbook

def nth_col_numbers(path, n, exclude=None, sheet_name=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active    
    nums = []
    rowCount = None
    if exclude is not None:
        rowCount = []
    i=0
    for row in ws.iter_rows(min_col=(n+1), max_col=(n+1), values_only=True):
        if exclude is not None:
            if i in exclude:
                i+=1
                continue
            i+=1

        val = row[0]
        if isinstance(val, (int, float)):
            nums.append(val)
        else:
            try:
                nums.append(float(val))
            except (TypeError, ValueError):
                pass
    return nums


path = sys.argv[1]

bubbleCount = nth_col_numbers(path,2)


binLabels = ["1", "2", "3", "4", "5+"]
binCounts = [0]*(len(binLabels))
for n in bubbleCount:
    try:
        iN = int(n)
    except Exception:
        continue
    if iN >= 5:
            binCounts[4] += 1
    else:
        binCounts[iN-1] += 1

binCountError = []
binCountError.append(np.sqrt(0))
ratios = [1]
ratioError = []
for c in binCounts[1:]:
    ratios.append(c/binCounts[0])
    binCountError.append(np.sqrt(c))    

for i in range(0,len(binLabels)):
    ratioError.append( np.sqrt(np.abs( (binCountError[i]/binCounts[0])**2 + (binCounts[i] * binCountError[0]/(binCounts[0]**2))**2   )))


## thresholds in eV, case B from ryan
thresholds = [0.0, 5000.0, 10000.0, 15000.0, 20000.0, 25000.0]
ratiosSim = []
simError  = []
ratiosSim.append([1,1,1,1,1,1])
simError.append([0,0,0,0,0,0])
ratiosSim.append([0.40904175, 0.3605948, 0.33270232,  0.31065533, 0.2928382,  0.26341731])
simError.append([0.01542234, 0.0195329, 0.0197859, 0.01941437, 0.01920887, 0.01817524])
ratiosSim.append([0.18451222, 0.12556795, 0.10411737, 0.08454227, 0.07374005, 0.06626506])
simError.append([0.00903129, 0.00975382, 0.00928462, 0.00839415, 0.00795298, 0.00757482])
ratiosSim.append([0.08392819, 0.04750103, 0.03265499, 0.02751376, 0.02068966, 0.01642935])
simError.append([0.00549518, 0.00539489, 0.00464159, 0.00432533, 0.00378953, 0.00338405])
ratiosSim.append([0.03417694, 0.01363073, 0.01183152, 0.00800400, 0.00636605, 0.00657174])
simError.append([0.00322163, 0.00264983, 0.00262369, 0.00218002, 0.00198435, 0.00205089])







## bubble finder accuracy check
## rows that i think we should exclude bc the event itself is weird
## reasoning is in note doc
eventsToIgnore = [ 27, 106, 115, 126, 145, 162, 178, 193, 199, 202 ]
## triplet, (eventIndex, hand count, sim count)
incorrectCount = []
excludedBinCounts = [0] * len(binLabels)
singlesCount = 0
singlesWrong = 0
multiCount = 0
multiWrong = 0

excludedCounts = nth_col_numbers(path, 2, eventsToIgnore)
bubbleFinderCount = nth_col_numbers(path,3,eventsToIgnore)

for i in range(0,len(bubbleFinderCount)-1):
    if excludedCounts[i] >= 5:
        excludedBinCounts[4] += 1
    else:
        excludedBinCounts[excludedCounts[i]-1] +=1
    if bubbleFinderCount[i] != excludedCounts[i]:
        incorrectCount.append((i,excludedCounts[i],bubbleFinderCount[i]))
        if excludedCounts[i] == 1:
            singlesWrong +=1
        else:
            multiWrong +=1

singlesCount = excludedBinCounts[0]
mutliCount = 0
for i in excludedBinCounts[1:]:
    multiCount += i


print("Amount of single bubble events miscounted:\t{}/{}\nAmount of multi-bubble events miscounted:\t{}/{}".format(singlesWrong,singlesCount,multiWrong,multiCount))
print("Total not ignored events:"+str(singlesCount+multiCount))


simCountMin = []
simCountMax = []
for i in range(0,len(thresholds)-1):
    simCountMin.append([])
    simCountMax.append([])
    for j in range(0,len(ratiosSim[i])-1):
        simCountMin[i].append(np.abs(singlesCount * (ratiosSim[j][i] - simError[j][i])))
        simCountMax[i].append(np.abs(singlesCount * (ratiosSim[j][i] + simError[j][i])))

x = np.arange(len(binLabels))

plt.figure(figsize=(16,9))
barsList = []
width = 0.9/(len(thresholds)-1)
colors = ["blue","red","green","orange", "teal","black"]
num_groups = len(thresholds) - 1

for i in range(num_groups):
    offset = (i - (num_groups - 1) / 2) * width
    xPos = x + offset
    bars = plt.bar(xPos, simCountMax[i], width=width, color=colors[i % len(colors)],
                   edgecolor="black", alpha=0.18,zorder=0)
    barsList.append(bars)


for i in range(num_groups):
    for j, bar in enumerate(barsList[i]):
        r = ratiosSim[j][i]    # note swapped indices
        cx = bar.get_x() + bar.get_width()/2
        cy = bar.get_height()
        plt.text(cx, cy + 0.01*max(binCounts), f"{r:0.3f}", ha='center', va='bottom', fontsize=12)

shadedInBars = []
for i in range(num_groups):
    offset = (i - (num_groups - 1) / 2) * width
    xPos = x + offset
    bars = plt.bar(xPos, simCountMin[i], width=width, color=colors[i % len(colors)],
                   edgecolor="black", label=f"{thresholds[i]/1000}KeV",zorder=2)
    shadedInBars.append(bars)




points = x
plt.errorbar(points, excludedBinCounts, yerr=binCountError,fmt='o',color="red", ecolor="red")


plt.xticks(x,binLabels)
plt.xlabel("Bubble Multiplicity",fontsize=16)
plt.ylabel("Count",fontsize=16)
plt.title("Handscan Bubble Multiplicites and Simulation Comparison",fontsize=16)
plt.legend(title="Thresholds",fontsize=16)
plt.savefig("multhist.png")
plt.show()



